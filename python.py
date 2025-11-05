# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import io, re, unicodedata
import requests
from datetime import date, datetime
from itertools import count
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

try:
    import pycountry
except Exception:
    pycountry = None

# ========== CONFIG & TITLE ==========
st.set_page_config(page_title="TẠO LỆNH CHUYỂN TIỀN QUỐC TẾ", page_icon="💸", layout="wide")
st.markdown(
    """
    <h1 style="text-align:center;color:#8B0000;">
        <span style="padding:6px 12px;border:2px solid #8B0000;border-radius:10px;">
            TẠO LỆNH CHUYỂN TIỀN QUỐC TẾ
        </span>
    </h1>
    """,
    unsafe_allow_html=True
)

# ========== HELPERS ==========
HTML_TAG_RE = re.compile(r"<[^>]+>")
NBSP = "\u00A0"

def parse_vn_number(s: str) -> float:
    if s is None: return 0.0
    s = str(s).strip().replace(NBSP, " ")
    s = HTML_TAG_RE.sub(" ", s)
    if s == "": return 0.0
    s = s.replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try: return float(s)
    except Exception: return 0.0

def fmt_vn_int(n): 
    try: return f"{int(round(float(n),0)):,}".replace(",", ".")
    except: return "0"

def fmt_usd(n):
    try: return f"{float(n):,.2f}"
    except: return "0.00"

def fmt_ddmmyyyy(d):
    if isinstance(d,(date,datetime)): return d.strftime("%d/%m/%Y")
    return ""

def vn_date_line(d: date, tinh="Lâm Đồng"):
    return f"{tinh}, ngày {d.day:02d} tháng {d.month:02d} năm {d.year}"

def clean_ccy(v)->str:
    if v is None: return ""
    s=str(v).strip().replace(NBSP," "); s=HTML_TAG_RE.sub(" ", s).upper()
    return s if re.fullmatch(r"[A-Z]{3}", s) else ""

def to_usd(amount, vnd_per_ccy, vnd_per_usd):
    if amount is None or pd.isna(amount): return 0.0
    if not (vnd_per_ccy and vnd_per_usd) or vnd_per_ccy<=0 or vnd_per_usd<=0: return 0.0
    return float(amount)*float(vnd_per_ccy)/float(vnd_per_usd)

# số nguyên → chữ (VN) đơn giản cho chứng từ
VN_NUM = ["không","một","hai","ba","bốn","năm","sáu","bảy","tám","chín"]
def _read_three(n):
    n = int(n)
    tr, du = divmod(n,100)
    ch, dv = divmod(du,10)
    s = ""
    if tr>0:
        s += VN_NUM[tr] + " trăm"
        if ch==0 and dv>0: s += " linh"
    if ch>1:
        s += " " + VN_NUM[ch] + " mươi"
        if dv==1: s += " mốt"
        elif dv==5: s += " lăm"
        elif dv>0: s += " " + VN_NUM[dv]
    elif ch==1:
        s += " mười"
        if dv==5: s += " lăm"
        elif dv>0: s += " " + VN_NUM[dv]
    else:
        if dv>0: s += " " + VN_NUM[dv]
    return s.strip()

def int_to_vn_words(n):
    n = int(round(float(n),0))
    if n==0: return "không"
    units = ["", " nghìn", " triệu", " tỷ", " nghìn tỷ", " triệu tỷ"]
    parts = []
    i=0
    while n>0 and i<len(units):
        n, r = divmod(n, 1000)
        if r>0:
            parts.append(_read_three(r) + units[i])
        i+=1
    return " ".join(reversed(parts)).strip()

def amount_to_words_vn(n, ccy=""):
    n = int(round(float(n),0))
    return f"{int_to_vn_words(n)} {ccy}".strip()

def get_country_list():
    items=[]
    if pycountry:
        try:
            for c in pycountry.countries:
                items.append((c.alpha_2.upper(), f"{c.alpha_2.upper()} – {c.name}"))
        except: pass
    if not items:
        fallback={"VN":"Viet Nam","US":"United States","AU":"Australia","JP":"Japan","KR":"Korea, Republic of",
                  "SG":"Singapore","CN":"China","DE":"Germany","FR":"France","GB":"United Kingdom","TH":"Thailand","CA":"Canada"}
        items=[(k,f"{k} – {v}") for k,v in fallback.items()]
    items.sort(key=lambda x:x[0]); return items

def get_currency_codes():
    codes=set()
    if pycountry:
        try:
            for c in pycountry.currencies:
                if getattr(c,"alpha_3",None): codes.add(c.alpha_3.upper())
        except: pass
    if not codes:
        codes={"USD","EUR","JPY","GBP","AUD","CAD","CHF","CNY","HKD","SGD","KRW","THB","VND"}
    return sorted(list(codes))

def names_loose_match(a: str, b: str) -> bool:
    def norm(s):
        if s is None: return []
        s=str(s).replace(NBSP," "); s=HTML_TAG_RE.sub(" ", s)
        s=unicodedata.normalize("NFKD", s)
        s="".join(ch for ch in s if not unicodedata.combining(ch))
        s=s.lower(); s=re.sub(r"[^a-z0-9]+", " ", s)
        toks=[t for t in s.split() if t]
        stop={"co","ltd","company","the","and","account","acc","fees","fee","university","bank",
              "beneficiary","name","accountname","transfer","payment","inv"}
        return [t for t in toks if t not in stop]
    A,B=set(norm(a)),set(norm(b))
    if not A or not B: return False
    if A.issubset(B) or B.issubset(A): return True
    inter=len(A&B); jacc=inter/max(1,len(A|B))
    return jacc>=0.7

def fetch_gdp_per_capita_usd(iso2: str, year: int):
    if not iso2 or not year: return None, None
    for y in [year, year-1, year-2]:
        try:
            u=f"https://api.worldbank.org/v2/country/{iso2.lower()}/indicator/NY.GDP.PCAP.CD?date={y}:{y}&format=json"
            js=requests.get(u,timeout=12).json()
            if isinstance(js,list) and len(js)>1 and js[1] and js[1][0]["value"] is not None:
                return float(js[1][0]["value"]), y
        except Exception:
            pass
    return None, None

# ========== UI HELPERS ==========
_key_counter = count(1)
def uk(prefix:str)->str: return f"{prefix}_{next(_key_counter)}"

def inline_input(label_text, widget_fn, *args, key_prefix=None, **kwargs):
    left, right = st.columns([0.38, 0.62])
    with left: st.markdown(f"**{label_text}**")
    with right:
        kwargs.setdefault("label_visibility","collapsed")
        if "key" not in kwargs:
            base = key_prefix or label_text.replace(" ","_").lower()
            kwargs["key"]=uk(base)
        return widget_fn("", *args, **kwargs)

# ========== 1. NGƯỜI CHUYỂN | 2. NGƯỜI NHẬN ==========
COUNTRIES = get_country_list()
COUNTRY_LABELS = [x[1] for x in COUNTRIES]
CURRENCIES = get_currency_codes()

left_col, right_col = st.columns(2)

with left_col:
    st.subheader("1. Người chuyển")
    send_date = inline_input("Ngày gửi tiền", st.date_input, value=date.today(), format="DD/MM/YYYY", key_prefix="send_date")
    pay_method = inline_input("Hình thức thanh toán", st.radio, options=["Tiền mặt","Chuyển khoản"], horizontal=True, index=0, key_prefix="pay_method")
    # Luôn hiển thị 3 ô tài khoản (có thể để trống)
    s_acc = inline_input("Số tài khoản (có thể để trống)", st.text_input, key_prefix="sender_acc")
    s_acc_name = inline_input("Tên tài khoản (có thể để trống)", st.text_input, key_prefix="sender_acc_name")
    s_acc_bank = inline_input("Tại ngân hàng (có thể để trống)", st.text_input, key_prefix="sender_acc_bank")

    s_full = inline_input("Họ tên", st.text_input, key_prefix="sender_full")
    s_addr = inline_input("Địa chỉ", st.text_area, height=80, key_prefix="sender_addr")
    s_country_label = inline_input("Quốc gia", st.selectbox, options=COUNTRY_LABELS,
                                   index=COUNTRY_LABELS.index("VN – Viet Nam") if "VN – Viet Nam" in COUNTRY_LABELS else 0,
                                   key_prefix="sender_country")
    s_country_code = s_country_label.split("–")[0].strip()
    s_country_name = s_country_label.split("–")[-1].strip()
    s_id_type = inline_input("Loại giấy tờ", st.selectbox, options=["CCCD","CC","Passport","Khác (tự nhập)"], index=0, key_prefix="sender_id_type")
    s_id_type_other = inline_input("Giấy tờ khác (nếu chọn Khác)", st.text_input, key_prefix="sender_id_type_other") if s_id_type=="Khác (tự nhập)" else ""
    s_id_no = inline_input("Số giấy tờ", st.text_input, key_prefix="sender_id_no")
    s_id_issue = inline_input("Ngày cấp", st.date_input, format="DD/MM/YYYY", key_prefix="sender_id_issue")
    s_id_place = inline_input("Nơi cấp giấy tờ", st.text_input, key_prefix="sender_id_place")
    s_phone = inline_input("Số điện thoại", st.text_input, key_prefix="sender_phone")

with right_col:
    st.subheader("2. Người nhận")
    r_full = inline_input("Họ tên", st.text_input, key_prefix="recv_full")
    r_acc = inline_input("Số tài khoản", st.text_input, key_prefix="recv_acc")
    r_addr = inline_input("Địa chỉ", st.text_area, height=80, key_prefix="recv_addr")
    r_cc_choice = inline_input("Mã quốc gia", st.selectbox, options=COUNTRY_LABELS,
                               index=COUNTRY_LABELS.index("VN – Viet Nam") if "VN – Viet Nam" in COUNTRY_LABELS else 0,
                               key_prefix="recv_cc")
    r_country_code = r_cc_choice.split("–")[0].strip()
    r_country_name = r_cc_choice.split("–")[-1].strip()
    r_id_type = inline_input("Loại giấy tờ (tuỳ chọn)", st.selectbox,
                             options=["(Để trống)","CCCD","CC","Passport","Khác (tự nhập)"], index=0, key_prefix="recv_id_type")
    r_id_type_other = inline_input("Giấy tờ khác (nếu chọn Khác)", st.text_input, key_prefix="recv_id_type_other") if r_id_type=="Khác (tự nhập)" else ""
    r_id_no = inline_input("Số giấy tờ (tuỳ chọn)", st.text_input, key_prefix="recv_id_no")

# ========== 3–6 ==========
secL, secR = st.columns(2)

with secL:
    st.subheader("3. Ngân hàng")
    inter_bank = inline_input("Ngân hàng trung gian", st.text_input, key_prefix="inter_bank")
    inter_swift = inline_input("SWIFT trung gian", st.text_input, key_prefix="inter_swift")
    ben_bank = inline_input("Ngân hàng nhận tiền", st.text_input, key_prefix="ben_bank")
    ben_swift = inline_input("SWIFT nhận tiền", st.text_input, key_prefix="ben_swift")

    st.subheader("4. Hồ sơ cung cấp")
    doc_opts=["CCCD","Giấy khai sinh","Passport","Visa","Thông báo học phí","Khác"]
    docs = inline_input("Chọn loại hồ sơ", st.multiselect, options=doc_opts, default=[], key_prefix="docs")
    doc_counts={}
    if docs:
        for d in docs:
            doc_counts[d] = inline_input(f"Số lượng '{d}'", st.number_input, min_value=1, value=1, step=1, key_prefix=f"doc_count_{d}")

with secR:
    st.subheader("5. Mục đích và số tiền")
    pay_type = inline_input("Loại thanh toán (Cá nhân)", st.selectbox, options=["Trợ cấp","Học phí","Mục đích khác"], index=0, key_prefix="pay_type")
    purpose_desc = inline_input("Nội dung chuyển tiền", st.text_area, height=80, key_prefix="purpose")
    currency = inline_input("Mã tiền tệ", st.selectbox, options=CURRENCIES,
                            index=CURRENCIES.index("USD") if "USD" in CURRENCIES else 0, key_prefix="currency")
    amt_str = inline_input("Số tiền ngoại tệ (VN: 1.234.567,89)", st.text_input, key_prefix="amt")
    vnd_per_ngt_str = inline_input("Tỷ giá VND/NGT (VND cho 1 NGT)", st.text_input, value="0", key_prefix="vnd_ngt")
    vnd_per_usd_str = inline_input("Tỷ giá VND/USD (VND cho 1 USD)", st.text_input, value="0", key_prefix="vnd_usd")
    fee_str = inline_input("Phí dịch vụ (VND)", st.text_input, value="0", key_prefix="fee")
    telex_str = inline_input("Điện phí (VND)", st.text_input, value="0", key_prefix="telex")

    foreign_amt = parse_vn_number(amt_str or "0")
    vnd_per_ngt = parse_vn_number(vnd_per_ngt_str or "0")
    vnd_per_usd = parse_vn_number(vnd_per_usd_str or "0")
    fee = parse_vn_number(fee_str or "0")
    telex = parse_vn_number(telex_str or "0")

    vnd_amount = round(foreign_amt * vnd_per_ngt, 0)
    total_vnd = vnd_amount + fee + telex
    usd_current = to_usd(foreign_amt, vnd_per_ngt, vnd_per_usd)

    st.markdown(
        f"""
        <div style="display:flex;gap:24px;flex-wrap:wrap;">
          <div style="flex:1;min-width:260px;background:#fafafa;padding:12px;border-radius:10px;">
            <div style="font-size:14px;color:#555;">Quy đổi (VND)</div>
            <div style="font-size:28px;font-weight:700;">{fmt_vn_int(vnd_amount)}</div>
          </div>
          <div style="flex:1;min-width:260px;background:#fafafa;padding:12px;border-radius:10px;">
            <div style="font-size:14px;color:#555;">Tổng thu (VND)</div>
            <div style="font-size:28px;font-weight:700;">{fmt_vn_int(total_vnd)}</div>
          </div>
          <div style="flex:1;min-width:260px;background:#fafafa;padding:12px;border-radius:10px;">
            <div style="font-size:14px;color:#555;">Giá trị hiện tại (USD)</div>
            <div style="font-size:28px;font-weight:700;">{fmt_usd(usd_current)}</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )

# ========== 6. LỊCH SỬ ==========
st.subheader("6. Lịch sử chuyển tiền")
st.markdown('<div style="color:#b00020;font-weight:700;">VUI LÒNG TẢI LÊN FILE .XLSX</div>', unsafe_allow_html=True)
hist_file = st.file_uploader("Tải file lịch sử (.xlsx)", type=["xlsx"], key=uk("hist_upload"))

def read_history_xlsx(uploaded_file) -> pd.DataFrame:
    empty = pd.DataFrame(columns=["recipient","amount","ccy","prepared date"])
    if uploaded_file is None: return empty
    df = pd.read_excel(uploaded_file)
    if df is None or df.empty: return empty

    def find_col(df, exact, contains=()):
        cols = {str(c).strip().lower(): c for c in df.columns}
        for k in exact:
            if k in cols: return cols[k]
        for k in list(exact)+list(contains):
            for ck, oc in cols.items():
                if k in ck: return oc
        return None

    recip = find_col(df, ["recipient","người nhận","nguoi nhan","beneficiary","receiver name","creditor name","account name"], ["beneficiar","receiver","creditor","account","name"])
    amt   = find_col(df, ["amount","số tiền","so tien","value","gia tri","amt"])
    ccy   = find_col(df, ["ccy","currency","mã tiền","ma tien","cur","tiền tệ"])
    dcol  = find_col(df, ["prepared date","value date","post date","posting date","transaction date","ngày","date"])

    out = pd.DataFrame()
    if recip is not None: out["recipient"] = df[recip].astype(str).str.strip()
    if amt   is not None: out["amount"]    = df[amt].apply(parse_vn_number).astype(float)
    if ccy   is not None: out["ccy"]       = df[ccy].apply(clean_ccy)
    else: out["ccy"] = ""
    if dcol  is not None: out["prepared date"] = pd.to_datetime(df[dcol], errors="coerce", dayfirst=True)
    else: out["prepared date"] = pd.NaT

    out = out[(out["recipient"].astype(str)!="") & (out["amount"].fillna(0)!=0)]
    return out.reset_index(drop=True)

hist_df = read_history_xlsx(hist_file)

# ========== KIỂM TRA HẠN MỨC ==========
st.markdown("---")
check_btn = st.button("✅ Kiểm tra hạn mức (GDP/người, quy đổi USD)", key=uk("check_btn")) if (pay_type=="Trợ cấp") else None

cap_usd=cap_year_used=remain_usd=None
summary_df=pd.DataFrame(columns=["Recipient","CCY","Amount_Total","Amount_Total_USD"])
total_usd_all=0.0
warning_text=""

if check_btn and (r_full or "").strip():
    cap_usd, cap_year_used = fetch_gdp_per_capita_usd(r_country_code, send_date.year) if r_country_code else (None, None)

    st.markdown(
        f"""
        <div style="margin:8px 0;padding:12px;background:#fff6e5;border:1px solid #ffe1b3;border-radius:10px;">
          <div style="font-size:16px;font-weight:600;">HẠN MỨC TRỢ CẤP (GDP/người):</div>
          <div style="font-size:30px;font-weight:800;color:#0b5;">{fmt_usd(cap_usd) if cap_usd is not None else 'Không lấy được dữ liệu'}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # Lọc theo người nhận; quy đổi USD dùng luôn tỷ giá ở Mục 5
    if not hist_df.empty:
        matched = hist_df[hist_df["recipient"].astype(str).apply(lambda x: names_loose_match(x, r_full))].copy()
    else:
        matched = pd.DataFrame()

    if not matched.empty:
        def row_to_usd(row):
            amt, ccy_row = row["amount"], row.get("ccy","")
            if ccy_row == "USD": 
                return float(amt) if pd.notna(amt) else 0.0
            # CCY trùng loại tiền đang chọn → dùng VND/NGT & VND/USD đã nhập ở Mục 5
            if ccy_row == currency:
                return to_usd(amt, vnd_per_ngt, vnd_per_usd)
            # CCY khác: không yêu cầu nhập lại, mặc định 0 (bỏ qua)
            return 0.0

        matched["Amount_Total_USD"] = matched.apply(row_to_usd, axis=1)
        summary_df = matched.groupby("ccy", dropna=False).agg(
            Amount_Total=("amount","sum"),
            Amount_Total_USD=("Amount_Total_USD","sum")
        ).reset_index().rename(columns={"ccy":"CCY"})
        summary_df["Recipient"]=r_full
        summary_df=summary_df[["Recipient","CCY","Amount_Total","Amount_Total_USD"]]
        total_usd_all=float(summary_df["Amount_Total_USD"].sum())

    st.markdown(
        f"""
        <div style="margin:8px 0;padding:12px;background:#eef7ff;border:1px solid #cfe6ff;border-radius:10px;">
          <div style="font-size:16px;font-weight:600;">SỐ TIỀN ĐÃ CHUYỂN (USD):</div>
          <div style="font-size:30px;font-weight:800;color:#0366d6;">{fmt_usd(total_usd_all)}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

    if cap_usd is not None:
        remain_usd = cap_usd - total_usd_all
        st.markdown(
            f"""
            <div style="margin:8px 0;padding:12px;background:#eafff1;border:1px solid #baf7d3;border-radius:10px;">
              <div style="font-size:16px;font-weight:600;">SỐ CÒN ĐƯỢC CHUYỂN (USD):</div>
              <div style="font-size:34px;font-weight:900;color:#0a7;">{fmt_usd(remain_usd)}</div>
            </div>
            """,
            unsafe_allow_html=True
        )
        if to_usd(foreign_amt, vnd_per_ngt, vnd_per_usd) > remain_usd or remain_usd < 0:
            st.markdown('<div style="font-size:34px;font-weight:900;color:#b00020;">🚨 BÁO ĐỘNG: CHUYỂN VƯỢT HẠN MỨC</div>', unsafe_allow_html=True)
            warning_text = "CHUYỂN VƯỢT HẠN MỨC"

    with st.expander("Bảng cộng dồn theo CCY (lọc đúng người nhận & quy đổi USD)", expanded=True):
        st.dataframe(summary_df, use_container_width=True)

# ========== XUẤT EXCEL (3 SHEET) ==========
st.markdown("---")
st.subheader("Xuất Excel")
template = st.file_uploader("Tải file Excel **mẫu in lệnh** (.xlsx)", type=["xlsx"], key=uk("template_upload"))

def compose_row_dict():
    items=[]
    for d in (docs or []):
        qty = int(st.session_state.get(f'doc_count_{d}',1))
        items.append(f"{qty} - {d}")  # theo cấu trúc yêu cầu
    docs_str="; ".join(items)

    return {
        "Ngày gửi": fmt_ddmmyyyy(send_date),
        "Hình thức thanh toán": pay_method,
        "Số tài khoản": s_acc,
        "Tên tài khoản": s_acc_name,
        "Tại ngân hàng": s_acc_bank,
        "Họ tên người chuyển": s_full,
        "Địa chỉ người chuyển": s_addr,
        "Quốc gia người chuyển": s_country_name,
        "Loại giấy tờ người chuyển": (s_id_type if s_id_type!="Khác (tự nhập)" else s_id_type_other),
        "Số giấy tờ người chuyển": s_id_no,
        "Ngày cấp GTTT người chuyển": fmt_ddmmyyyy(s_id_issue),
        "Nơi cấp GTTT người chuyển": s_id_place,
        "SĐT người chuyển": s_phone,
        "Họ tên người nhận": r_full,
        "Số tài khoản người nhận": r_acc,
        "Địa chỉ người nhận": r_addr,
        "Quốc gia người nhận": r_country_name,
        "Ngân hàng trung gian": inter_bank,
        "SWIFT trung gian": inter_swift,
        "Ngân hàng nhận tiền": ben_bank,
        "SWIFT nhận tiền": ben_swift,
        "Loại thanh toán (Cá nhân)": pay_type,
        "Nội dung chuyển tiền": purpose_desc,
        "Hồ sơ cung cấp": docs_str,
        "Mã tiền tệ": currency,
        "Số tiền ngoại tệ": parse_vn_number(amt_str or "0"),
        "Tỷ giá VND/NGT": parse_vn_number(vnd_per_ngt_str or "0"),
        "Tỷ giá VND/USD": parse_vn_number(vnd_per_usd_str or "0"),
        "Số tiền quy đổi (VND)": int(round(parse_vn_number(amt_str or "0")*parse_vn_number(vnd_per_ngt_str or "0"),0)),
        "Phí dịch vụ (VND)": int(round(parse_vn_number(fee_str or "0"),0)),
        "Điện phí (VND)": int(round(parse_vn_number(telex_str or "0"),0)),
        "Tổng thu (VND)": int(round(parse_vn_number(amt_str or "0")*parse_vn_number(vnd_per_ngt_str or "0") + parse_vn_number(fee_str or "0") + parse_vn_number(telex_str or "0"),0)),
        "Giá trị giao dịch hiện tại (USD)": to_usd(parse_vn_number(amt_str or "0"), parse_vn_number(vnd_per_ngt_str or "0"), parse_vn_number(vnd_per_usd_str or "0")),
        "Hạn mức (USD)": cap_usd if cap_usd is not None else "",
        "Đã chuyển (USD)": total_usd_all,
        "Còn được chuyển (USD)": (cap_usd - total_usd_all) if cap_usd is not None else "",
        "Cảnh báo": warning_text or "Được chuyển",
    }

def bold_tnr(cell, value):
    cell.value = value
    cell.font = Font(name="Times New Roman", bold=True)

def fill_command_sheet(ws, data):
    amt_words = amount_to_words_vn(data["Số tiền ngoại tệ"], data["Mã tiền tệ"])
    vnd_words = amount_to_words_vn(data["Số tiền quy đổi (VND)"], "đồng")

    bold_tnr(ws["E11"], data["Ngày gửi"])
    bold_tnr(ws["I11"], f'{data["Mã tiền tệ"]} {int(round(data["Số tiền ngoại tệ"],0)):,}'.replace(",", "."))
    bold_tnr(ws["G14"], amt_words)

    bold_tnr(ws["J15"], data["Số tài khoản"])
    bold_tnr(ws["H16"], data["Số giấy tờ người chuyển"])
    bold_tnr(ws["K16"], data["Loại giấy tờ người chuyển"])

    bold_tnr(ws["H18"], data["Ngày cấp GTTT người chuyển"])
    bold_tnr(ws["K18"], data["Nơi cấp GTTT người chuyển"])

    bold_tnr(ws["A18"], data["Họ tên người chuyển"])
    bold_tnr(ws["A20"], f'{data["Địa chỉ người chuyển"]}, {data["Quốc gia người chuyển"]}')
    bold_tnr(ws["H19"], data["SĐT người chuyển"])

    bold_tnr(ws["G21"], data["Ngân hàng trung gian"])
    bold_tnr(ws["D22"], data["SWIFT trung gian"])
    bold_tnr(ws["G23"], data["Ngân hàng nhận tiền"])
    bold_tnr(ws["D24"], data["SWIFT nhận tiền"])

    bold_tnr(ws["A27"], data["Họ tên người nhận"])
    bold_tnr(ws["H27"], data["Số tài khoản người nhận"])
    bold_tnr(ws["A29"], f'{data["Địa chỉ người nhận"]}, {data["Quốc gia người nhận"]}')
    bold_tnr(ws["A31"], data["Nội dung chuyển tiền"])

    bold_tnr(ws["B39"], "x" if data["Hình thức thanh toán"]=="Tiền mặt" else "")
    bold_tnr(ws["B40"], "x" if data["Hình thức thanh toán"]=="Chuyển khoản" else "")
    bold_tnr(ws["F40"], data["Số tài khoản"] if data["Hình thức thanh toán"]=="Chuyển khoản" else "")
    bold_tnr(ws["J40"], data["Tại ngân hàng"] if data["Hình thức thanh toán"]=="Chuyển khoản" else "")

    bold_tnr(ws["A50"], data["Hồ sơ cung cấp"] or "")

    bold_tnr(ws["H66"], vn_date_line(send_date))
    bold_tnr(ws["F75"], data["Mã tiền tệ"])
    bold_tnr(ws["D76"], f'{int(round(data["Số tiền ngoại tệ"],0)):,} {data["Mã tiền tệ"]}'.replace(",", "."))
    bold_tnr(ws["D77"], amt_words)

    bold_tnr(ws["F83"], data["Tỷ giá VND/NGT"])
    bold_tnr(ws["H83"], f'VNĐ/{data["Mã tiền tệ"]}')

    # C86 để trống, C87 là số tiền VND bằng chữ
    bold_tnr(ws["D86"], int(round(data["Số tiền quy đổi (VND)"],0)))
    ws["C86"].value = None
    bold_tnr(ws["C87"], vnd_words)

    bold_tnr(ws["H94"], vn_date_line(send_date))

def export_excel(template_file, mapping: dict, summary: pd.DataFrame, warnings: str) -> bytes:
    if template_file is None:
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            pd.DataFrame([{"Lưu ý":"Chưa có file mẫu. Sheet này chỉ minh hoạ."}]).to_excel(w, index=False, sheet_name="LỆNH CHUYỂN TIỀN")
            pd.DataFrame([mapping]).to_excel(w, index=False, sheet_name="THÔNG TIN CHUYỂN TIỀN")
            warn_df = pd.DataFrame([{
                "Tên người nhận": mapping["Họ tên người nhận"],
                "Hạn mức": mapping.get("Hạn mức (USD)",""),
                "Số tiền đã chuyển": mapping.get("Đã chuyển (USD)",""),
                "Số tiền lần này": mapping.get("Giá trị giao dịch hiện tại (USD)",""),
                "Cảnh báo": warnings or "Được chuyển",
            }])
            warn_df.to_excel(w, index=False, sheet_name="CẢNH BÁO")
        out.seek(0); return out.read()

    bio = io.BytesIO(template_file.read()); wb = load_workbook(bio)
    ws = wb.active
    fill_command_sheet(ws, mapping)

    # Sheet THÔNG TIN CHUYỂN TIỀN
    if "THÔNG TIN CHUYỂN TIỀN" in wb.sheetnames: wb.remove(wb["THÔNG TIN CHUYỂN TIỀN"])
    ws_info = wb.create_sheet("THÔNG TIN CHUYỂN TIỀN")
    df_info = pd.DataFrame([mapping])
    for r in dataframe_to_rows(df_info, index=False, header=True): ws_info.append(r)
    # Đặt font đậm + TNR cho các ô giá trị (hàng 2 trở đi)
    for row in ws_info.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name="Times New Roman", bold=True)

    # Sheet CẢNH BÁO
    if "CẢNH BÁO" in wb.sheetnames: wb.remove(wb["CẢNH BÁO"])
    ws_warn = wb.create_sheet("CẢNH BÁO")
    headers = ["Tên người nhận","Hạn mức","Số tiền đã chuyển","Số tiền lần này","Cảnh báo"]
    ws_warn.append(headers)
    alert_text = warnings or "Được chuyển"
    row = [
        mapping["Họ tên người nhận"],
        mapping.get("Hạn mức (USD)",""),
        mapping.get("Đã chuyển (USD)",""),
        mapping.get("Giá trị giao dịch hiện tại (USD)",""),
        alert_text,
    ]
    ws_warn.append(row)
    # bôi đậm đỏ cột cảnh báo (hàng 2, cột 5)
    warn_cell = ws_warn["E2"]
    warn_cell.font = Font(name="Times New Roman", bold=True, color="FF0000")

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()

row_dict = compose_row_dict()
excel_bytes = export_excel(template, row_dict, summary_df, warning_text)
st.download_button(
    "⬇️ Tải file Excel (3 sheet: LỆNH CHUYỂN TIỀN / THÔNG TIN CHUYỂN TIỀN / CẢNH BÁO)",
    data=excel_bytes,
    file_name=f"lenh_chuyen_tien_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    key=uk("download_btn")
)
